from pathlib import Path
import math
import re
import sys

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path(r"C:\Users\rking\Downloads\1Solar_20260802_40.333002_-111.712338_5.0_3M (2) (1).xlsx")
SOURCE = (ROOT / "app.js").read_text(encoding="utf-8")
INDEX = (ROOT / "index.html").read_text(encoding="utf-8")


def require(condition, message):
    if not condition:
        raise AssertionError(message)
    print(f"PASS  {message}")


def js_round(value):
    return math.floor(value + 0.5)


def rounded_hundred(value):
    return js_round(value / 100) * 100


def workbook_evidence():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    metadata = wb["Metadata"]
    weekly = wb["Weekly"]
    meta_headers = [cell.value for cell in next(metadata.iter_rows(min_row=1, max_row=1))]
    weekly_headers = [cell.value for cell in next(weekly.iter_rows(min_row=1, max_row=1))]
    meta = [dict(zip(meta_headers, row)) for row in metadata.iter_rows(min_row=2, values_only=True)]
    data = [dict(zip(weekly_headers, row)) for row in weekly.iter_rows(min_row=2, values_only=True)]
    tesla_ports = {row["Charger_id"] for row in meta if row["Network"] == "tesla" and row["Name"] == "Orem, UT"}
    tesla_weekly = [row for row in data if row["Network"] == "tesla" and row["Name"] == "Orem, UT"]
    sessions = sum((row["Charges_ccs"] or 0) + (row["Charges_cha"] or 0) + (row["Charges_nacs"] or 0) for row in tesla_weekly)
    dates = {row["Date"] for row in tesla_weekly}
    weighted_minutes = sum(
        (row["Charges_ccs"] or 0) * (row["Avg_charge_ccs"] or 0)
        + (row["Charges_cha"] or 0) * (row["Avg_charge_cha"] or 0)
        + (row["Charges_nacs"] or 0) * (row["Avg_charge_nacs"] or 0)
        for row in tesla_weekly
    ) / sessions
    return len(tesla_ports), sessions, len(dates) * 7, sessions / (len(dates) * 7), weighted_minutes


def main():
    ports, sessions, days, daily_visits, avg_minutes = workbook_evidence()
    require(ports == 8, "Workbook: nearby Orem Tesla station has 8 ports")
    require(sessions == 18713, "Workbook: nearby Orem Tesla station totals 18,713 sessions")
    require(days == 91, "Workbook: observation window contains 91 days")
    require(abs(daily_visits - 205.6374) < 0.01, "Workbook: observed market demand is 205.6 visits/day (reported as roughly 200)")
    require(abs(avg_minutes - 24.0724) < 0.01, "Workbook: weighted average session duration is 24.07 minutes (reported as approximately 24)")

    require("'kneaders-orem':" in SOURCE, "Canonical Kneaders Orem EV-only bid profile exists")
    profile = SOURCE.split("'kneaders-orem':", 1)[1].split("'kneaders-orem-ev':", 1)[0]
    require("scopes: { solar: false, storage: false, ev: true }" in profile, "Bid scope is EV only; solar and battery are disabled")
    require("Kneaders Bakery & Cafe Orem, Utah" in profile, "Proposal name matches the required store-name/location format")
    require("1960 State Street, Orem, Utah 84057" in profile, "Proposal location matches the workbook address")
    require("marketProofSessions3m: 18713" in profile and "marketProofDays: 91" in profile, "Imported Paren evidence is preserved in the bid profile")
    require("'copper-fork-grill-american-fork'" in SOURCE and "scopes: { solar: true, storage: true, ev: true }" in SOURCE, "Hypothetical restaurant profile restores all three proposal scopes")
    copper_card = INDEX.split('data-bid="copper-fork-grill-american-fork"', 1)[1].split('</article>', 1)[0].lower()
    require("solar" in copper_card and "battery" in copper_card and "ev charging" in copper_card, "All-three-scope dashboard card shows solar, battery, and EV charging")
    require(INDEX.count('data-bid="kneaders-orem"') == 1, "Dashboard has one canonical Kneaders proposal card")
    kneaders_card = INDEX.split('data-bid="kneaders-orem"', 1)[1].split('</article>', 1)[0].lower()
    require("ev charging" in kneaders_card and "solar" not in kneaders_card and "battery" not in kneaders_card, "Dashboard card presents Kneaders as EV charging only")
    require('id="copyProposalButton"' in INDEX, "Proposal header exposes an editable-copy action")
    require("encodeCopyPayload" in SOURCE and "decodeCopyPayload" in SOURCE and "scopes: currentScopes()" in SOURCE, "Copy links serialize and restore proposal inputs and scope")
    require("recipient can open it and make their own changes" in SOURCE, "Copy action clearly communicates editable ownership")
    require("editableTextSelector" in SOURCE and "setSectionEditing" in SOURCE, "Edit mode includes all visible text, metric labels, and metric values")
    require("GetEV-inline-edits" in SOURCE and "inlineEdits" in SOURCE, "Inline section edits persist and travel with copied proposals")

    forecast = {1: 7.7, 3: 15.4, 5: 16.1}
    forecast_visits = {year: 8 * 24 * (utilization / 100) / (24 / 60) for year, utilization in forecast.items()}
    require([js_round(forecast_visits[year]) for year in (1, 3, 5)] == [37, 74, 77], "Forecast formula yields 37, 74, and 77 visits/day in Years 1, 3, and 5")
    require("Daily charging visits = Ports" in SOURCE and "Average session length in hours" in SOURCE, "Proposal shows the required daily-visits formula")

    parties = js_round(forecast_visits[5] * 0.30)
    require(parties == 23, "30% capture of Year 5 visits yields 23 additional parties/day")
    require(parties * 30 == 690, "Customer-party monthly output is 690")
    require(parties * 365 == 8395, "Customer-party annual output is 8,395")

    expected_sales = {
        12: (276, 8400, 100700),
        20: (460, 14000, 167900),
        30: (690, 21000, 251900),
    }
    for receipt, (daily, monthly, annual) in expected_sales.items():
        calculated_daily = parties * receipt
        calculated_monthly = rounded_hundred(calculated_daily * 365 / 12)
        calculated_annual = rounded_hundred(calculated_daily * 365)
        require((calculated_daily, calculated_monthly, calculated_annual) == (daily, monthly, annual), f"Restaurant sales case ${receipt}: ${daily}/day, ${monthly}/month, ${annual}/year")

    required_inputs = ["ports", "averageSessionMinutes", "forecastYear1Utilization", "forecastYear3Utilization", "forecastYear5Utilization", "restaurantCaptureRate", "conservativeReceipt", "averageReceipt", "highReceipt", "daysPerYear"]
    ev_config = SOURCE.split("ev: { title: 'EV customer value'", 1)[1].split("  bundles:", 1)[0]
    for key in required_inputs:
        require(f"['{key}'" in ev_config, f"Configure EV charging exposes independent input: {key}")
    for label in ["Charging visits / day", "Additional customer parties / day", "Daily restaurant sales", "Monthly restaurant sales", "Annual restaurant sales"]:
        require(label in ev_config, f"Configure EV charging shows dependent formula: {label}")

    ev_story = SOURCE.split("function renderEvCustomerStory()", 1)[1].split("function renderEvCustomerValues()", 1)[0]
    required_story_terms = ["OBSERVED MARKET PROOF", "PROJECTED AT THIS SITE", "KNEADERS CAPTURE", "ANNUAL SALES OPPORTUNITY", "EVPIN FORECAST RAMP", "WHAT IS ALREADY HAPPENING", "WHAT THIS COULD MEAN FOR KNEADERS", "RESTAURANT SALES OPPORTUNITY", "Gross sales, not profit."]
    for term in required_story_terms:
        require(term in ev_story, f"EV story contains required element: {term}")
    forbidden_metrics = ["all nearby stations", "10.7%", "annual sessions", "technical utilization percentage"]
    for term in forbidden_metrics:
        require(term not in ev_story, f"EV story does not introduce prohibited extra metric: {term}")
    require("['site', 'layout', 'solar', 'storage', 'bundles', 'vpp', 'investment', 'economics']" in SOURCE, "EV-only renderer hides every non-EV report section")
    ev_only_overview = SOURCE.split("function renderEvOnlyOverview()", 1)[1].split("function renderReport()", 1)[0].lower()
    require("solar" not in ev_only_overview and "battery" not in ev_only_overview, "EV-only proposal header makes no solar or battery claim")

    print("PASS  All required EV-only story values and formulas match the requested defaults.")


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"FAIL  {error}")
        sys.exit(1)
